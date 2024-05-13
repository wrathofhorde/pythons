import openpyxl as op
import os
import schedule
import time

#엑셀파일, 시트 정보를 읽어서 리스트로 리턴
def sheetInfo(path : str) -> list:

    excellist = os.listdir(path)
    shtlist = []
    for file in excellist:
        temp_wb = op.load_workbook(path+"/"+file)
        temp_shtlist = temp_wb.sheetnames
        for sht in temp_shtlist:
            temp_tuple = (file, sht)
            shtlist.append(temp_tuple)
    
    return shtlist

#path는 엑셀 파일이 있는 경로, shtlist는 sheetInfo 함수의 리턴값임
def hyperLink(path, shtlist : list):
    
    wb = op.Workbook() #엑셀 워크북 신규 생성
    ws = wb.active #엑셀 활성화시트 설정(신규 워크북이므로 Sheet1)
    
    i=1 #엑셀파일에 데이터를 쓰기 위한 인덱싱

    for data in shtlist:
        ws.cell(row=i+1, column=1).value = data[0] #튜플 0번째 요소 : 파일명
        ws.cell(row=i+1, column=2).value = data[1] #튜플 1번째 요소 : 시트명

        #엑셀 Hyperlink 함수 구현
        ws.cell(row=i+1, column=3).value = '=HYPERLINK("{}", "{}")'.format("["+path+"/"+data[0]+"]"+data[1]+"!A1","링크")
        i=i+1
    
    ws["A1"].value = "파일명"
    ws["B1"].value = "시트명"
    ws["C1"].value = "하이퍼링크"
    wb.save("시트목록.xlsx")


#실행
# path = r"엑셀파일 경로"
# sht_info = sheetInfo(path)
# hyperLink(path,sht_info)


#엑셀파일, 시트 정보를 읽어서 리스트로 리턴
def getSheetInfo(file : str) -> list:

    shtlist = []
    temp_wb = op.load_workbook(file)
    temp_shtlist = temp_wb.sheetnames
    
    for sht in temp_shtlist:
        shtlist.append(sht)
    
    return shtlist

#path는 엑셀 파일이 있는 경로, shtlist는 sheetInfo 함수의 리턴값임
def getHyperLink(file, shtlist : list):
    ns = "목록"
    wb = op.load_workbook(file)
    ws = wb.create_sheet(ns, 0) #1번째 자리에 목록 시트 생성
    # ws = wb.active #엑셀 활성화시트 설정(신규 워크북이므로 Sheet1)

    
    i = 2 #엑셀파일에 데이터를 쓰기 위한 인덱싱, 타이틀을 위해 2부터 시작

    for sht in shtlist:
        # ws.cell(row=i+1, column=1).value = '=HYPERLINK("{}", "{}")'.format(sht+"!A1",sht)
        ws.cell(row = i, column = 1).value = i - 1
        ws.cell(row = i, column = 2).value = sht
        ws.cell(row = i, column = 2).hyperlink = "#"+ sht + "!A1"
        ws.cell(row = i, column = 2).style = "Hyperlink"
        i = i + 1
    
    ws["A1"].value = "번호"
    ws["B1"].value = "시트명"
    wb.save(file)
    wb.close()

file = r"pravang_main_db_테이블명세_240513.xlsx"
sht_info = getSheetInfo(file)
# print(sht_info)
getHyperLink(file, sht_info)