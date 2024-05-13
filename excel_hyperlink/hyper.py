import openpyxl as op
import os

#엑셀파일, 시트 정보를 읽어서 리스트로 리턴
def getSheetInfo(file : str) -> list:

    shtlist = []
    temp_wb = op.load_workbook(file)
    temp_shtlist = temp_wb.sheetnames
    
    for sht in temp_shtlist:
        shtlist.append(sht)
    
    return shtlist

#file 엑셀 파일 이름, shtlist는 getSheetInfo 함수의 리턴값임
def getHyperLink(file, shtlist : list):
    ns = "목록"
    wb = op.load_workbook(file)
    ws = wb.create_sheet(ns, 0) #1번째 자리에 목록 시트 생성
    
    i = 2 #엑셀파일에 데이터를 쓰기 위한 인덱싱, 타이틀을 위해 2부터 시작

    for sht in shtlist:
        ws.cell(row = i, column = 1).value = i - 1 #시트 번호
        ws.cell(row = i, column = 2).value = sht
        ws.cell(row = i, column = 2).hyperlink = "#"+ sht + "!A1" # 엑셀 파일에서 현재 문서는 #
        ws.cell(row = i, column = 2).style = "Hyperlink"
        i = i + 1
    
    ws["A1"].value = "번호"
    ws["B1"].value = "시트명"
    wb.save(file)
    wb.close()

file = r"table.xlsx"
sht_info = getSheetInfo(file)
# print(sht_info)
getHyperLink(file, sht_info)